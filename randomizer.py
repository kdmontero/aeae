from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import random

# use this when reading workbook thru the script is available
def execute() -> None:
    NUMBER_OF_CATEGORIES = 62

    wb = load_workbook('aeae_categories.xlsx')
    ws = wb.worksheets[0]

    categories = []
    database_count = []
    questions_count = []

    for i in range(2, NUMBER_OF_CATEGORIES + 2):
        categories.append(ws[f'A{i}'].value)
        database_count.append(ws[f'B{i}'].value)
        questions_count.append(ws[f'C{i}'].value)

    random_questions_indices = []
    current_index = 1

    for required, total in zip(questions_count, database_count):
        indices = sorted(random.sample(
            range(current_index, current_index + total), required
        ))

        random_questions_indices += indices
        current_index += total

    print(random_questions_indices)
    random.shuffle(random_questions_indices)
    print(random_questions_indices)

    
# use this when executing the script in browser and reading a workbook is not possible
def execute_simple_script() -> None:
    NUMBER_OF_CATEGORIES = 62

    categories = [] # Place here the list of categories for reference only. Not required in running the script
    database_count = [] # Place here the total number of questions in database per categories (list of numbers)
    questions_count = [] # Place here the required number of questions (list of numbers)

    random_questions_indices = []
    current_index = 1

    for required, total in zip(questions_count, database_count):
        indices = sorted(random.sample(
            range(current_index, current_index + total), required
        ))

        random_questions_indices += indices
        current_index += total

    print(random_questions_indices)
    random.shuffle(random_questions_indices)
    print(random_questions_indices)
    
    
if __name__ == '__main__':
    execute()
